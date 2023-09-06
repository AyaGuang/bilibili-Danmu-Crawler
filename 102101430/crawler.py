import requests
import re
import itertools
import sys
import time
from fake_useragent import UserAgent
from multiprocessing import Pool
from collections import Counter
from openpyxl import Workbook

#定义一些参数
uA = UserAgent()
target = "日本核污染水排海"
total_Num = 300 #爬取300个视频、请保证能被step整除
step = 30 #每次爬取30个视频
BVs = [] #存储bv号
counts = Counter()
header = {#从浏览器中复制下请求头
        'Accept':'application/json, text/plain, */*',
        'Accept-Encoding':'gzip, deflate, br',
        'Accept-Language':'zh-CN,zh;q=0.9',
        'Cookie':'i-wanna-go-back=-1; buvid4=B4C50C10-9B84-829C-550D-419C5FA4126D63704-022012116-tKxIwyBtfce6CxQ0ElglAw%3D%3D; CURRENT_BLACKGAP=0; buvid_fp_plain=undefined; LIVE_BUVID=AUTO5016571134973707; fingerprint3=c5c219316d0ccb004373ae81c1fccd50; rpdid=|(JlRYJ~Yk)k0J\'uYY)l~kJlm; hit-new-style-dyn=1; CURRENT_PID=01c3a640-cd15-11ed-be2f-b9d6f8fcd3c4; FEED_LIVE_VERSION=V8; buvid3=69F5EB42-E69A-FE9E-7403-89217946A89D91721infoc; b_nut=1688646492; _uuid=F38573DB-92FE-685E-E4110-CDD12936C28495998infoc; hit-dyn-v2=1; CURRENT_QUALITY=80; b_ut=5; header_theme_version=CLOSE; fingerprint=347a82351e94d7a58ecc0fb8a9c7eff7; CURRENT_FNVAL=4048; home_feed_column=5; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2OTQwOTY5MDYsImlhdCI6MTY5MzgzNzcwNiwicGx0IjotMX0.z334MNCOGLvZKPDOo41gKoOsiwLXDtcen17hEnloQ-E; bili_ticket_expires=1694096906; browser_resolution=1707-917; buvid_fp=347a82351e94d7a58ecc0fb8a9c7eff7; SESSDATA=959e95a3%2C1709474725%2Cfa205%2A9247T3X6UVfRmt2Q14XTH_gosu6lH1BvPuuVlKcbVJv64tO74rYwvS8zlT8XnN0KmFRweMJAAAGwA; bili_jct=08c90b43a4f60a67a47cf85922b0a8eb; DedeUserID=65993927; DedeUserID__ckMd5=a4ecada1426e7447; sid=73zq6pro; bp_video_offset_65993927=837879399983349907; PVID=4; b_lsid=3FBAC57F_18A6612D649',
        'Origin':'https://search.bilibili.com',
        'Referer':'https://search.bilibili.com/all?vt=17031316&keyword=%E6%97%A5%E6%9C%AC%E6%A0%B8%E6%B1%A1%E6%9F%93%E6%B0%B4%E6%8E%92%E6%B5%B7&page=2&o=24',
        'Sec-Ch-Ua':'"Chromium";v="116", "Not)A;Brand";v="24", "Google Chrome";v="116"',
        'Sec-Ch-Ua-Mobile':'?0',
        'Sec-Ch-Ua-Platform':'"Windows"',
        'Sec-Fetch-Dest':'empty',
        'Sec-Fetch-Mode':'cors',
        'Sec-Fetch-Site':'same-site',
        'User-Agent':uA.random
}

header1 = {
    'UserAgent' : uA.random
}

def get_Response(html_Url,h = header):  # 对于任意一个网址url，获取其html文件
    
    response = requests.get(html_Url, headers=h)
    if response.ok:
        response.encoding = 'utf-8'
        return response
    # 如果执行到这一步，就说明请求失败，报错并退出
    print(f"请求网页失败,状态码：{response.status_code}")
    sys.exit()


def get_Target_Video(page): # 从B站搜索栏中找排名前total_Num的视频，爬取他们的BV号，target为搜索内容,step为一页读取的视频数，默认为30
    params = {  #针对搜索栏，从浏览器中复制参数调用api
        '__refresh__': 'true',
        '_extra':'' ,
        'context':'' ,
        'page': page,
        'page_size': step,#设置一页读取step个视频
        'from_source': '',
        'from_spmid': '333.337',
        'platform': 'pc',
        'highlight': '1',
        'single_column': '0',
        'keyword': target,#此处是搜索内容
        'qv_id': '00d1q7iUbrvfsPdFZv2zXXtDqfcWuzER',
        'ad_resource': '5654',
        'source_tag': '3',
        'gaia_vtoken': '',
        'category_id': '',
        'search_type': 'video',
        'dynamic_offset': (step+1)*(page-1),
        'web_location': '1430654',
        'w_rid': '47e55164f507eec457da5802b3ea5d70',
        'wts': '1693922979'
    }
    api_Url = "https://api.bilibili.com/x/web-interface/wbi/search/type"#搜索栏的api
    print(f"开始爬取第{page}页视频")
    content = requests.get(api_Url,params,headers= header).text
    #从格式为 "bvid":"BV1194y1z7M3"中提取BV号
    match_Text = r'"bvid":"([^"]+)"'
    print(f"已爬取第{page}页视频")
    return re.findall(match_Text, content)


def get_Cid(BV):  # 对于一个B站视频网址，获取其cid(即存储弹幕的网站key),输入为BV号
    # print(f"正在解析{BV}的cid")
    bili_Url = f"https://www.ibilibili.com/video/{BV}/"
    # print(bili_Url)
    content = get_Response(bili_Url).text
    # 我们用正则从html文件中提取 "bcid":"1252950136" 中的cid号
    match_Text = r'"bcid":"(\d*)"'
    match = re.search(match_Text, content)
    if match:
        cid = match.group(1)
        return cid
    # 如果执行到这一步，说明没有匹配到cid，报错并退出
    print("解析弹幕cid失败")
    sys.exit()


def get_Danmaku(BV):  # 对于一个BV号，将其分析为cid，后进入弹幕网站并爬取弹幕进行分析
    cid = get_Cid(BV)
    # print(f"正在解析cid:{cid}的弹幕")
    # danmaku(だんまく)就是弹幕的意思
    danmaku_Url = f"https://api.bilibili.com/x/v1/dm/list.so?oid={cid}"
    content = get_Response(danmaku_Url,header1).text
    # 这样写看上去和上一个函数没什么两样，但我觉得可读性应该会好一点
    # 这次的格式是<d p=......>我是弹幕</d>
    match_Text = r'<d[^>]*>([^<]+)</d>'  
    matches = re.findall(match_Text, content)#爬取到的弹幕全部存储在列表中
    new_Counts = Counter(matches)#利用counter库对列表中的弹幕计数
    return new_Counts


def main():
    # Step 1：利用线程池找到前综合排序前300个视频
    print("Step 1: 利用线程池找到前综合排序前300个视频")
    pool = Pool(10)
    bv = pool.map(get_Target_Video,range(1,(total_Num//step)+1))#得到一个list，其每个元素为装着bv号的list
    BVs = list(itertools.chain(*bv))#将得到的列表整合到一起
    pool.close()
    pool.join()
    # print(BVs)


    # Step 2：利用线程池对每一个视频的弹幕进行分析
    print("Step 2: 利用线程池对每一个视频的弹幕进行分析")
    pool = Pool(6)  # 线程太多会403、412，少用一些线程
    all_counts = pool.map(get_Danmaku,BVs)
    for count in all_counts:
        counts.update(count)
    print("分析完成,正在导入Excel....")
    pool.close()
    pool.join()

    # Step 3: 取出现次数前20的弹幕，将他导入Excel
    print("Step 3: 取出现次数前20的弹幕，将他导入Excel")
    Top_20_Count = counts.most_common(20)
    all_counts = counts.most_common()
    # 创建一个新的 Excel 工作簿
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "Top 20"
    # 将出现次数前 20 多的元素和计数器写入 Excel 表格中
    for row, (item, count) in enumerate(Top_20_Count, start=1):
        worksheet1.cell(row=row, column=1, value=item)
        worksheet1.cell(row=row, column=2, value=count)

    # 在第二页保存所有弹幕留档
    worksheet2 = workbook.create_sheet(title="所有弹幕")
    for row, (item, count) in enumerate(all_counts, start=1):
        worksheet2.cell(row=row, column=1, value=item)
        worksheet2.cell(row=row, column=2, value=count)
    # 保存 Excel 工作簿
    workbook.save('output/Top_20_danmu.xlsx')
    print("导入完成！")


if __name__ == "__main__":
    main()
    # input("Press <Enter> to close\n")
